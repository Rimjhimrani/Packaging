import streamlit as st
import pandas as pd
import numpy as np
import os
import json
import hashlib
import nltk
from datetime import datetime
from difflib import SequenceMatcher
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import re
import io
import tempfile
import shutil
from pathlib import Path

# Configure Streamlit page
st.set_page_config(
    page_title="AI Template Mapper",
    page_icon="🤖",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Try to import optional dependencies
try:
    from nltk.tokenize import word_tokenize
    from nltk.corpus import stopwords
    from sklearn.feature_extraction.text import TfidfVectorizer
    from sklearn.metrics.pairwise import cosine_similarity
    
    # Initialize NLTK with better error handling
    def initialize_nltk():
        """Initialize NLTK with proper downloads and fallbacks"""
        try:
            # Try to download required NLTK data
            required_downloads = [
                ('punkt', 'tokenizers/punkt'),
                ('punkt_tab', 'tokenizers/punkt_tab'), 
                ('stopwords', 'corpora/stopwords')
            ]
            
            for download_name, path in required_downloads:
                try:
                    nltk.data.find(path)
                except LookupError:
                    try:
                        nltk.download(download_name, quiet=True)
                    except Exception as e:
                        print(f"Warning: Could not download {download_name}: {e}")
            
            # Test tokenization
            word_tokenize("test")
            return True
            
        except Exception as e:
            print(f"NLTK initialization failed: {e}")
            return False
    
    # Initialize NLTK
    NLTK_READY = initialize_nltk()
    
    if NLTK_READY:
        ADVANCED_NLP = True
    else:
        # Fallback: disable NLTK features if initialization fails
        ADVANCED_NLP = False
        st.warning("⚠️ NLTK initialization failed. Using basic text processing.")
        
except ImportError as e:
    ADVANCED_NLP = False
    NLTK_READY = False
    st.warning("⚠️ Advanced NLP features disabled. Install nltk and scikit-learn for better matching.")

class AdvancedTemplateMapper:
    def __init__(self):
        self.similarity_threshold = 0.3
        self.stop_words = {
            'a', 'an', 'and', 'are', 'as', 'at', 'be', 'by', 'for', 'from',
            'has', 'he', 'in', 'is', 'it', 'its', 'of', 'on', 'that', 'the',
            'to', 'was', 'will', 'with', 'or', 'but', 'not', 'this', 'have',
            'had', 'what', 'when', 'where', 'who', 'which', 'why', 'how'
        }
        
        if ADVANCED_NLP:
            try:
                self.stop_words = set(stopwords.words('english'))
                self.vectorizer = TfidfVectorizer(stop_words='english', ngram_range=(1, 2))
            except:
                pass
        
    def preprocess_text(self, text):
        """Preprocess text for better matching"""
        try:
            if pd.isna(text) or text is None:
                return ""
            
            text = str(text).lower()
            text = re.sub(r'[^\w\s]', ' ', text)
            text = re.sub(r'\s+', ' ', text).strip()
            
            return text
        except Exception as e:
            st.error(f"Error in preprocess_text: {e}")
            return ""
    
    def extract_keywords(self, text):
        """Extract keywords from text with improved error handling"""
        try:
            text = self.preprocess_text(text)
            if not text:
                return []
            # Try NLTK tokenization if available
            if ADVANCED_NLP and NLTK_READY:
                try:
                    tokens = word_tokenize(text)
                    keywords = [token for token in tokens if token not in self.stop_words and len(token) > 2]
                    return keywords
                except Exception as e:
                    # If NLTK fails, fall back to simple tokenization
                    print(f"NLTK tokenization failed, using fallback: {e}")
            # Fallback: Simple tokenization
            tokens = text.split()
            keywords = [token for token in tokens if token not in self.stop_words and len(token) > 2]
            return keywords
        except Exception as e:
            st.error(f"Error in extract_keywords: {e}")
            return []
            
    def simple_tokenize(text):
        """Simple tokenization without NLTK dependency"""
        # Remove punctuation and split
        text = re.sub(r'[^\w\s]', ' ', text.lower())
        tokens = text.split()
        return [token for token in tokens if len(token) > 2]
    
    def calculate_similarity(self, text1, text2):
        """Calculate similarity between two texts"""
        try:
            if not text1 or not text2:
                return 0.0
            
            text1 = self.preprocess_text(text1)
            text2 = self.preprocess_text(text2)
            
            if not text1 or not text2:
                return 0.0
            
            # Sequence similarity
            sequence_sim = SequenceMatcher(None, text1, text2).ratio()
            
            # TF-IDF similarity (if available)
            tfidf_sim = 0.0
            if ADVANCED_NLP:
                try:
                    tfidf_matrix = self.vectorizer.fit_transform([text1, text2])
                    tfidf_sim = cosine_similarity(tfidf_matrix[0:1], tfidf_matrix[1:2])[0][0]
                except:
                    tfidf_sim = 0.0
            
            # Keyword overlap
            keywords1 = set(self.extract_keywords(text1))
            keywords2 = set(self.extract_keywords(text2))
            
            if keywords1 and keywords2:
                keyword_sim = len(keywords1.intersection(keywords2)) / len(keywords1.union(keywords2))
            else:
                keyword_sim = 0.0
            
            # Weighted average
            if ADVANCED_NLP:
                final_similarity = (sequence_sim * 0.4) + (tfidf_sim * 0.4) + (keyword_sim * 0.2)
            else:
                final_similarity = (sequence_sim * 0.7) + (keyword_sim * 0.3)
            
            return final_similarity
        except Exception as e:
            st.error(f"Error in calculate_similarity: {e}")
            return 0.0
    
    def is_data_cell(self, cell_value):
        """Determine if a cell is meant for data entry"""
        try:
            if not cell_value or pd.isna(cell_value):
                return True
            
            cell_str = str(cell_value).strip()
            if not cell_str:
                return True
            
            # Data placeholder patterns
            data_patterns = [
                r'^_+$', r'^\.*$', r'^-+$', r'^\[.*\]$', r'^\{.*\}$', r'^<.*>$',
                r'enter|fill|data|value|input|here|placeholder', r'^\d{1,2}/\d{1,2}/\d{2,4}$',
                r'^dd/mm/yyyy|mm/dd/yyyy|yyyy-mm-dd$', r'^\$\d*\.?\d*$', r'^\d*\.?\d*$',
            ]
            
            cell_lower = cell_str.lower()
            
            for pattern in data_patterns:
                if re.search(pattern, cell_lower):
                    return True
            
            # Special character dominated cells
            if len(cell_str) <= 10 and len(re.sub(r'[a-zA-Z0-9]', '', cell_str)) > len(cell_str) * 0.5:
                return True
            
            return False
        except Exception as e:
            st.error(f"Error in is_data_cell: {e}")
            return False
    
    def is_section_header(self, text):
        """Identify section headers that should never be mapped"""
        try:
            if not text or pd.isna(text):
                return False
                
            text = str(text).strip()
            if not text:
                return False
                
            text_lower = text.lower()
            
            section_patterns = [
                'packaging instruction', 'vendor information', 'part information', 'current packaging',
                'primary packaging', 'secondary packaging', 'packaging procedure', 'reference image',
                'problem', 'instruction', 'details', 'specification', 'requirements', 'process',
                'procedure', 'approved by', 'reviewed by', 'issued by'
            ]
            
            for pattern in section_patterns:
                if pattern in text_lower:
                    return True
            
            if len(text.split()) > 3 and not text.endswith(':') and len(text) > 15:
                return True
                
            return False
        except Exception as e:
            st.error(f"Error in is_section_header: {e}")
            return False
    
    def is_table_header(self, text):
        """Identify table headers that should never be mapped"""
        try:
            if not text or pd.isna(text):
                return False
                
            text = str(text).strip()
            if not text:
                return False
                
            text_lower = text.lower()
            
            table_patterns = [
                'l-mm', 'w-mm', 'h-mm', 'length', 'width', 'height', 'dimension',
                'qty/pack', 'pack weight', 'empty weight', 'total', 'packaging type',
                'weight', 'quantity', 'size', 'volume', 'capacity'
            ]
            
            for pattern in table_patterns:
                if pattern in text_lower:
                    return True
            
            if re.search(r'mm|cm|kg|gm|pcs|qty|pack|dimension|weight|size', text_lower):
                return True
                
            return False
        except Exception as e:
            st.error(f"Error in is_table_header: {e}")
            return False
    
    def is_label_cell(self, text):
        """Identify mappable field labels"""
        try:
            if not text or pd.isna(text):
                return False
                
            text = str(text).strip()
            if not text:
                return False
            
            if (self.is_data_cell(text) or 
                self.is_section_header(text) or 
                self.is_table_header(text)):
                return False
            
            mappable_fields = [
                'code', 'name', 'part no', 'description', 'revision no', 'revision',
                'vendor', 'supplier', 'customer', 'client', 'company', 'manufacturer',
                'address', 'phone', 'email', 'contact', 'reference', 'ref',
                'date', 'time', 'invoice', 'bill', 'order', 'id', 'number',
                'serial', 'batch', 'lot', 'model', 'version', 'type', 'category', 'Lmm','Wmm',
                'Hmm', 'Unit Weight', 'L- mm', 'W- mm', 'H- mm','Qty / Pack', 'Qty/Pack'
            ]
            
            text_lower = text.lower()
            
            for field in mappable_fields:
                if field in text_lower:
                    return True
            
            if text.endswith(':'):
                return True
            
            if (len(text.split()) <= 3 and 
                len(text) > 1 and 
                not text.isdigit() and
                not text.isupper() and
                len(text) < 20):
                return True
            
            return False
        except Exception as e:
            st.error(f"Error in is_label_cell: {e}")
            return False
    
    def classify_cell_type(self, cell_value):
        """Classify the cell type based on its content"""
        try:
            if not cell_value or pd.isna(cell_value):
                return 'data_cell'
            
            text = str(cell_value).strip()
            
            if not text:
                return 'data_cell'
            
            text_lower = text.lower()
            
            if self.is_data_cell(text):
                return 'data_cell'
            
            if self.is_section_header(text):
                return 'section_header'
            
            if self.is_table_header(text):
                return 'table_header'
            
            if len(text) > 50 or (text.isupper() and len(text.split()) >= 4):
                return 'title'
            
            field_keywords = ['name', 'number', 'date', 'time', 'code', 'id', 'description',
                              'weight', 'size', 'quantity', 'address', 'phone', 'email',
                              'vendor', 'customer', 'amount', 'price', 'total', 'type', 'part',
                              'reference', 'ref', 'model', 'version', 'serial', 'batch']
            
            if any(kw in text_lower for kw in field_keywords):
                return 'field_header'
            
            if text.endswith(':'):
                return 'field_header'
            
            if self.is_label_cell(text):
                return 'field_header'
            
            return 'data_cell'
            
        except Exception as e:
            st.error(f"Error in classify_cell_type: {e}")
            return 'data_cell'
    
    def find_template_fields(self, template_file):
        """Find all template fields with automatic classification"""
        fields = {}
        
        try:
            workbook = openpyxl.load_workbook(template_file)
            worksheet = workbook.active
            
            merged_ranges = worksheet.merged_cells.ranges
            
            for row in worksheet.iter_rows():
                for cell in row:
                    try:
                        if cell.value is not None:
                            cell_value = str(cell.value).strip()
                            
                            if cell_value:
                                cell_coord = cell.coordinate
                                merged_range = None
                                
                                for merge_range in merged_ranges:
                                    if cell.coordinate in merge_range:
                                        merged_range = str(merge_range)
                                        break
                                
                                cell_type = self.classify_cell_type(cell_value)
                                
                                fields[cell_coord] = {
                                    'value': cell_value,
                                    'row': cell.row,
                                    'column': cell.column,
                                    'merged_range': merged_range,
                                    'is_label': cell_type == 'field_header',
                                    'is_data_cell': cell_type == 'data_cell',
                                    'cell_type': cell_type
                                }
                    except Exception as e:
                        st.error(f"Error processing cell {cell.coordinate}: {e}")
                        continue
            
            workbook.close()
            
        except Exception as e:
            st.error(f"Error reading template: {e}")
        
        return fields
    
    def map_data_to_template(self, template_fields, data_df):
        """Automatically map data columns to template fields"""
        mapping_results = {}
        try:
            data_columns = data_df.columns.tolist()
            
            mappable_fields = {coord: field for coord, field in template_fields.items()
                               if field.get('cell_type') == 'field_header' or field.get('is_label') == True}
            
            for coord, field in mappable_fields.items():
                try:
                    best_match = None
                    best_score = 0.0
                    
                    for data_col in data_columns:
                        similarity = self.calculate_similarity(field['value'], data_col)
                        
                        if similarity > best_score and similarity >= self.similarity_threshold:
                            best_score = similarity
                            best_match = data_col
                    
                    mapping_results[coord] = {
                        'template_field': field['value'],
                        'data_column': best_match,
                        'similarity': best_score,
                        'field_info': field,
                        'is_mappable': best_match is not None
                    }
                        
                except Exception as e:
                    st.error(f"Error mapping field {coord}: {e}")
                    continue
                    
        except Exception as e:
            st.error(f"Error in map_data_to_template: {e}")
            
        return mapping_results
    
    def find_data_cell_for_label(self, worksheet, field_info):
        """Automatically find data cell for a label (improved merged cell handling)"""
        try:
            row = field_info['row']
            col = field_info['column']
            # Get merged ranges for reference
            merged_ranges = list(worksheet.merged_cells.ranges)
        
            def is_suitable_data_cell(cell_coord):
                """Check if a cell is suitable for data entry"""
                try:
                    cell = worksheet[cell_coord]
                    # Skip MergedCell objects (they're read-only)
                    if hasattr(cell, '__class__') and cell.__class__.__name__ == 'MergedCell':
                        return False
                    # Check if it's a data cell or empty
                    if cell.value is None or self.is_data_cell(cell.value):
                        return True
                    return False
                except:
                    return False
            # Strategy 1: Look right of label (most common pattern)
            for offset in range(1, 6):
                target_col = col + offset
                if target_col <= worksheet.max_column:
                    cell_coord = worksheet.cell(row=row, column=target_col).coordinate
                    if is_suitable_data_cell(cell_coord):
                        return cell_coord
            # Strategy 2: Look below label
            for offset in range(1, 4):
                target_row = row + offset
                if target_row <= worksheet.max_row:
                    cell_coord = worksheet.cell(row=target_row, column=col).coordinate
                    if is_suitable_data_cell(cell_coord):
                        return cell_coord
            # Strategy 3: Look in nearby area (diagonal search)
            for r_offset in range(-1, 3):
                for c_offset in range(-1, 6):
                    if r_offset == 0 and c_offset == 0:
                        continue
                    target_row = row + r_offset
                    target_col = col + c_offset
                
                    if (target_row > 0 and target_row <= worksheet.max_row and 
                        target_col > 0 and target_col <= worksheet.max_column):
                            cell_coord = worksheet.cell(row=target_row, column=target_col).coordinate
                            if is_suitable_data_cell(cell_coord):
                                return cell_coord
            # Strategy 4: If label is in a merged cell, try to find data cell in the same merged range
            if field_info.get('merged_range'):
                try:
                    for merged_range in merged_ranges:
                        label_coord = worksheet.cell(row=row, column=col).coordinate
                        if label_coord in merged_range:
                            # Look for empty cells within or adjacent to the merged range
                            min_row, min_col, max_row, max_col = merged_range.bounds
                        
                            # Check cells within the merged range
                            for r in range(min_row, max_row + 1):
                                for c in range(min_col, max_col + 1):
                                    cell_coord = worksheet.cell(row=r, column=c).coordinate
                                    if is_suitable_data_cell(cell_coord):
                                        return cell_coord
                            # Check cells adjacent to the merged range
                            for c in range(max_col + 1, max_col + 4):
                                if c <= worksheet.max_column:
                                    for r in range(min_row, max_row + 1):
                                        cell_coord = worksheet.cell(row=r, column=c).coordinate
                                        if is_suitable_data_cell(cell_coord):
                                            return cell_coord
                            break
                        
                except Exception as e:
                    st.warning(f"Error processing merged range for {field_info.get('value', 'unknown')}: {e}")
            return None
            
        except Exception as e:
            st.error(f"Error in find_data_cell_for_label: {e}")
            return None
    
    def fill_template_with_data(self, template_file, mapping_results, data_df):
        """Fill template with mapped data and return the filled workbook"""
        try:
            workbook = openpyxl.load_workbook(template_file)
            worksheet = workbook.active
            
            filled_count = 0
            
            for coord, mapping in mapping_results.items():
                try:
                    if mapping['data_column'] is not None and mapping['is_mappable']:
                        field_info = mapping['field_info']
                        
                        target_cell = self.find_data_cell_for_label(worksheet, field_info)
                        
                        if target_cell and len(data_df) > 0:
                            data_value = data_df.iloc[0][mapping['data_column']]
                            
                            cell_obj = worksheet[target_cell]
                            if hasattr(cell_obj, '__class__') and cell_obj.__class__.__name__ == 'MergedCell':
                                # Get top-left anchor of merged range
                                for merged_range in worksheet.merged_cells.ranges:
                                    if target_cell in merged_range:
                                        anchor_cell = merged_range.start_cell
                                        anchor_cell.value = str(data_value) if not pd.isna(data_value) else ""
                                        break
                            else:
                                cell_obj.value = str(data_value) if not pd.isna(data_value) else ""
                            filled_count += 1
                            
                except Exception as e:
                    st.error(f"Error filling mapping {coord}: {e}")
                    continue
            
            return workbook, filled_count
            
        except Exception as e:
            st.error(f"Error filling template: {e}")
            return None, 0

# Initialize session state
if 'authenticated' not in st.session_state:
    st.session_state.authenticated = False
if 'user_role' not in st.session_state:
    st.session_state.user_role = None
if 'templates' not in st.session_state:
    st.session_state.templates = {}
if 'ai_mapper' not in st.session_state:
    st.session_state.ai_mapper = AdvancedTemplateMapper()

# User management functions
def hash_password(password):
    return hashlib.sha256(password.encode()).hexdigest()

def verify_password(password, hashed):
    return hash_password(password) == hashed

# Default users
DEFAULT_USERS = {
    "admin": {
        "password": hash_password("admin123"),
        "role": "admin",
        "name": "Administrator"
    },
    "user1": {
        "password": hash_password("user123"),
        "role": "user",
        "name": "Regular User"
    }
}

def authenticate_user(username, password):
    if username in DEFAULT_USERS:
        if verify_password(password, DEFAULT_USERS[username]['password']):
            return DEFAULT_USERS[username]['role'], DEFAULT_USERS[username]['name']
    return None, None

# Login function
def show_login():
    st.title("🤖 AI Template Mapper")
    st.markdown("### Enhanced template processing with merged cell support")
    
    col1, col2, col3 = st.columns([1, 2, 1])
    
    with col2:
        with st.form("login_form"):
            st.subheader("Login")
            username = st.text_input("Username")
            password = st.text_input("Password", type="password")
            submit = st.form_submit_button("Login", use_container_width=True)
            
            if submit:
                role, name = authenticate_user(username, password)
                if role:
                    st.session_state.authenticated = True
                    st.session_state.user_role = role
                    st.session_state.username = username
                    st.session_state.name = name
                    st.rerun()
                else:
                    st.error("Invalid credentials")
        
        st.info("**Demo Credentials:**\n- Admin: admin / admin123\n- User: user1 / user123")

# Main dashboard
def show_dashboard():
    # Header
    col1, col2 = st.columns([3, 1])
    with col1:
        st.title(f"Welcome, {st.session_state.name} ({st.session_state.user_role})")
    with col2:
        if st.button("Logout", type="secondary"):
            st.session_state.authenticated = False
            st.session_state.user_role = None
            st.rerun()
    
    # Sidebar navigation
    with st.sidebar:
        st.header("Navigation")
        
        if st.session_state.user_role == 'admin':
            page = st.selectbox(
                "Select Page",
                ["Dashboard", "Upload Template", "View Templates", "Analyze Template", "AI Data Processor"]
            )
        else:
            page = st.selectbox(
                "Select Page", 
                ["Dashboard", "AI Data Processor", "View Templates"]
            )
    
    # Page routing
    if page == "Dashboard":
        show_dashboard_content()
    elif page == "Upload Template":
        show_upload_template()
    elif page == "View Templates":
        show_templates()
    elif page == "Analyze Template":
        show_analyze_template()
    elif page == "AI Data Processor":
        show_data_processor()

def show_dashboard_content():
    st.header("🚀 Enhanced AI Template System")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.subheader("✨ New Features")
        st.markdown("""
        - ✅ **Merged Cell Support**: Handles complex templates with merged cells
        - 🤖 **Advanced AI Mapping**: Uses NLP for intelligent field detection
        - 📊 **Automatic Form Recognition**: Detects fillable fields automatically
        - 🔍 **Template Analysis**: Comprehensive structure analysis
        - 🎯 **Smart Data Placement**: Intelligent data positioning
        """)
    
    with col2:
        st.subheader("📊 System Statistics")
        total_templates = len(st.session_state.templates)
        threshold = st.session_state.ai_mapper.similarity_threshold
        
        st.metric("Available Templates", total_templates)
        st.metric("AI Similarity Threshold", f"{threshold:.2f}")
        st.metric("Advanced Processing", "Active" if ADVANCED_NLP else "Basic")

def show_upload_template():
    if st.session_state.user_role != 'admin':
        st.error("Access denied")
        return
    
    st.header("📁 Upload Template")
    st.info("Upload any Excel template - even complex forms with merged cells! AI will automatically detect fillable fields.")
    
    with st.form("upload_form"):
        template_name = st.text_input("Template Name", placeholder="Enter template name")
        uploaded_file = st.file_uploader("Select Excel Template File", type=['xlsx'])
        submit = st.form_submit_button("Upload & Analyze Template")
        
        if submit and uploaded_file and template_name:
            try:
                with st.spinner("Analyzing template..."):
                    # Save uploaded file temporarily
                    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
                        tmp_file.write(uploaded_file.getvalue())
                        tmp_path = tmp_file.name
                    
                    # Analyze template
                    template_fields = st.session_state.ai_mapper.find_template_fields(tmp_path)
                    
                    # Determine template type
                    template_type = "Complex Form" if len(template_fields) > 10 else "Standard"
                    
                    # Store template data
                    st.session_state.templates[template_name] = {
                        'file_data': uploaded_file.getvalue(),
                        'fields': template_fields,
                        'field_count': len(template_fields),
                        'type': template_type,
                        'created_date': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        'created_by': st.session_state.username
                    }
                    
                    # Clean up temp file
                    os.unlink(tmp_path)
                    
                    st.success(f"Template '{template_name}' uploaded successfully!")
                    st.info(f"Detected {len(template_fields)} fields | Type: {template_type}")
                    
                    # Show field breakdown
                    field_types = {}
                    for field in template_fields.values():
                        cell_type = field.get('cell_type', 'unknown')
                        field_types[cell_type] = field_types.get(cell_type, 0) + 1
                    
                    st.subheader("Field Analysis")
                    for cell_type, count in field_types.items():
                        st.write(f"- {cell_type.replace('_', ' ').title()}: {count}")
                    
            except Exception as e:
                st.error(f"Failed to upload template: {str(e)}")

def show_templates():
    st.header("📋 Available Templates")
    
    if not st.session_state.templates:
        st.info("No templates available")
        return
    
    for template_name, template_info in st.session_state.templates.items():
        with st.expander(f"📋 {template_name}"):
            col1, col2 = st.columns(2)
            
            with col1:
                st.write(f"**Created:** {template_info.get('created_date', 'Unknown')}")
                st.write(f"**By:** {template_info.get('created_by', 'Unknown')}")
                st.write(f"**Type:** {template_info.get('type', 'Standard')}")
                
            with col2:
                st.write(f"**Fields:** {template_info.get('field_count', 0)}")
                
                if st.session_state.user_role == 'admin':
                    if st.button(f"Delete {template_name}", key=f"del_{template_name}"):
                        del st.session_state.templates[template_name]
                        st.rerun()

def show_analyze_template():
    if st.session_state.user_role != 'admin':
        st.error("Access denied")
        return
    
    st.header("🔍 Analyze Template")
    
    uploaded_file = st.file_uploader("Select Excel file to analyze", type=['xlsx'])
    
    if uploaded_file:
        try:
            with st.spinner("Analyzing template structure..."):
                with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
                    tmp_file.write(uploaded_file.getvalue())
                    tmp_path = tmp_file.name
                
                template_fields = st.session_state.ai_mapper.find_template_fields(tmp_path)
                os.unlink(tmp_path)
            
            st.success(f"Analysis complete! Found {len(template_fields)} fields")
            
            # Field breakdown
            field_types = {}
            for field in template_fields.values():
                cell_type = field.get('cell_type', 'unknown')
                if cell_type not in field_types:
                    field_types[cell_type] = []
                field_types[cell_type].append(field)
            
            # Display by type
            for cell_type, fields in field_types.items():
                with st.expander(f"{cell_type.replace('_', ' ').title()} ({len(fields)} fields)"):
                    for field in fields[:10]:  # Show first 10
                        st.write(f"• **{field['value']}** (Row {field['row']}, Col {field['column']})")
                        if field.get('merged_range'):
                            st.write(f"  └─ Merged range: {field['merged_range']}")
                    
                    if len(fields) > 10:
                        st.write(f"... and {len(fields) - 10} more")
                        
        except Exception as e:
            st.error(f"Error analyzing template: {str(e)}")

def show_data_processor():
    st.header("🤖 AI Data Processor")
    st.info("Upload your data file and select a template. AI will automatically map and fill the template!")
    
    # Data file upload
    data_file = st.file_uploader("Upload Data File", type=['csv', 'xlsx'])
    
    # Template selection
    if st.session_state.templates:
        selected_template = st.selectbox(
            "Select Template",
            options=list(st.session_state.templates.keys()),
            format_func=lambda x: f"{x} ({st.session_state.templates[x]['type']})"
        )
    else:
        st.warning("No templates available. Please upload a template first.")
        return
    
    if data_file and selected_template:
        try:
            # Load data
            if data_file.name.lower().endswith('.csv'):
                data_df = pd.read_csv(data_file)
            else:
                data_df = pd.read_excel(data_file)
            
            st.subheader("📊 Data Preview")
            st.dataframe(data_df.head(), use_container_width=True)
            
            if st.button("🚀 Process with AI", type="primary"):
                with st.spinner("🤖 AI is processing your data..."):
                    # Get template info
                    template_info = st.session_state.templates[selected_template]
                    template_fields = template_info['fields']
                    
                    # Create temporary template file
                    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
                        tmp_file.write(template_info['file_data'])
                        template_path = tmp_file.name
                    
                    # AI mapping
                    mapping_results = st.session_state.ai_mapper.map_data_to_template(template_fields, data_df)
                    
                    # Fill template
                    filled_workbook, filled_count = st.session_state.ai_mapper.fill_template_with_data(
                        template_path, mapping_results, data_df
                    )
                    
                    # Clean up temp file
                    os.unlink(template_path)
                
                if filled_workbook:
                    st.success(f"✅ Processing complete! Filled {filled_count} fields automatically.")
                    
                    # Show mapping results
                    st.subheader("🎯 AI Mapping Results")
                    
                    mapped_fields = [m for m in mapping_results.values() if m['is_mappable']]
                    unmapped_fields = [m for m in mapping_results.values() if not m['is_mappable']]
                    
                    col1, col2 = st.columns(2)
                    
                    with col1:
                        st.metric("Successfully Mapped", len(mapped_fields))
                        if mapped_fields:
                            st.write("**Mapped Fields:**")
                            for mapping in mapped_fields[:5]:  # Show first 5
                                confidence = mapping['similarity'] * 100
                                st.write(f"• {mapping['template_field']} ← {mapping['data_column']} ({confidence:.1f}%)")
                            if len(mapped_fields) > 5:
                                st.write(f"... and {len(mapped_fields) - 5} more")
                    
                    with col2:
                        st.metric("Unmapped Fields", len(unmapped_fields))
                        if unmapped_fields:
                            st.write("**Unmapped Fields:**")
                            for mapping in unmapped_fields[:5]:  # Show first 5
                                st.write(f"• {mapping['template_field']}")
                            if len(unmapped_fields) > 5:
                                st.write(f"... and {len(unmapped_fields) - 5} more")
                    
                    # Download filled template
                    st.subheader("📥 Download Results")
                    
                    # Save workbook to bytes
                    output = io.BytesIO()
                    filled_workbook.save(output)
                    output.seek(0)
                    
                    # Generate filename
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    filename = f"{selected_template}_filled_{timestamp}.xlsx"
                    
                    st.download_button(
                        label="📁 Download Filled Template",
                        data=output.getvalue(),
                        file_name=filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        type="primary"
                    )
                    
                    # Process multiple rows option
                    if len(data_df) > 1:
                        st.subheader("🔄 Batch Processing")
                        st.info(f"Your data has {len(data_df)} rows. Process all rows?")
                        
                        if st.button("🚀 Process All Rows", type="secondary"):
                            with st.spinner("Processing all data rows..."):
                                # Create zip file for multiple templates
                                zip_buffer = io.BytesIO()
                                
                                with tempfile.TemporaryDirectory() as temp_dir:
                                    filled_files = []
                                    
                                    for idx, row in data_df.iterrows():
                                        # Create single-row dataframe
                                        single_row_df = pd.DataFrame([row])
                                        
                                        # Create temp template file
                                        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
                                            tmp_file.write(template_info['file_data'])
                                            temp_template_path = tmp_file.name
                                        
                                        # Fill template for this row
                                        row_workbook, _ = st.session_state.ai_mapper.fill_template_with_data(
                                            temp_template_path, mapping_results, single_row_df
                                        )
                                        
                                        if row_workbook:
                                            # Save to temp directory
                                            row_filename = f"{selected_template}_row_{idx+1}_{timestamp}.xlsx"
                                            row_filepath = os.path.join(temp_dir, row_filename)
                                            row_workbook.save(row_filepath)
                                            filled_files.append((row_filename, row_filepath))
                                        
                                        # Clean up temp template file
                                        os.unlink(temp_template_path)
                                    
                                    # Create zip file
                                    import zipfile
                                    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                                        for filename, filepath in filled_files:
                                            zip_file.write(filepath, filename)
                                
                                zip_buffer.seek(0)
                                
                                st.success(f"✅ Processed {len(filled_files)} templates successfully!")
                                
                                st.download_button(
                                    label="📦 Download All Filled Templates (ZIP)",
                                    data=zip_buffer.getvalue(),
                                    file_name=f"{selected_template}_batch_{timestamp}.zip",
                                    mime="application/zip",
                                    type="primary"
                                )
                else:
                    st.error("❌ Failed to process template. Please check your data and template.")
                    
        except Exception as e:
            st.error(f"Error processing data: {str(e)}")
            st.exception(e)

# Configuration sidebar
def show_config_sidebar():
    with st.sidebar:
        st.header("⚙️ AI Configuration")
        
        # Similarity threshold
        new_threshold = st.slider(
            "Similarity Threshold",
            min_value=0.1,
            max_value=0.9,
            value=st.session_state.ai_mapper.similarity_threshold,
            step=0.05,
            help="Higher values require closer matches"
        )
        
        if new_threshold != st.session_state.ai_mapper.similarity_threshold:
            st.session_state.ai_mapper.similarity_threshold = new_threshold
            st.success("Threshold updated!")
        
        # Advanced settings
        st.subheader("🔧 Advanced Settings")
        
        st.info(f"**NLP Status:** {'Advanced' if ADVANCED_NLP else 'Basic'}")
        
        if not ADVANCED_NLP:
            st.warning("Install nltk and scikit-learn for better AI matching")
        
        # System info
        st.subheader("📊 System Info")
        st.write(f"Templates: {len(st.session_state.templates)}")
        st.write(f"User: {st.session_state.get('name', 'Unknown')}")
        st.write(f"Role: {st.session_state.get('user_role', 'Unknown')}")

# Main application
def main():
    if not st.session_state.authenticated:
        show_login()
    else:
        show_config_sidebar()
        show_dashboard()

if __name__ == "__main__":
    main()
